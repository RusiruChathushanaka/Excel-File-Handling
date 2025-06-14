"""
xlsx_utils.py - Python library for working with modern Excel (.xlsx) files
Requires: openpyxl
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Note: The helper functions col_idx_to_letters and letters_to_col_idx
# from your original script are available in openpyxl.utils, so we
# import and use them directly for consistency and reliability.

# --------------------------
# Core Functions
# --------------------------

def get_xlsx_cell_value(file_path, sheet_name, cell_ref):
    """
    Returns the value of a cell in an .xlsx Excel file.
    
    Args:
        file_path (str): Path to the .xlsx file.
        sheet_name (str): Name of the worksheet.
        cell_ref (str): Cell reference (e.g., 'C6').
    
    Returns:
        The value of the cell, or an error message if not found.
    """
    try:
        # Open the workbook and worksheet
        wb = openpyxl.load_workbook(file_path, data_only=True) # data_only=True to get values instead of formulas
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."
        ws = wb[sheet_name]
        
        # openpyxl can directly access cells by reference
        cell = ws[cell_ref]
        return cell.value
    except KeyError:
        return "Error: Invalid cell reference format or cell does not exist."
    except Exception as e:
        return f"Error: {str(e)}"

def update_xlsx_cell(file_path, sheet_name, cell_ref, update_value):
    """
    Updates a cell value in an .xlsx file.
    
    Args:
        file_path (str): Path to the .xlsx file.
        sheet_name (str): Name of the worksheet.
        cell_ref (str): Cell reference (e.g., 'C5').
        update_value: Value to set in the cell.
    
    Returns:
        str: 'Success' or error reason.
    """
    # Validate file existence
    if not os.path.isfile(file_path):
        return f"Failed: File not found - {file_path}"
    
    # Validate file extension
    if not file_path.lower().endswith('.xlsx'):
        return "Failed: Only .xlsx files are supported."
    
    try:
        # Open the workbook
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return f"Failed: Sheet '{sheet_name}' not found."
        
        ws = wb[sheet_name]
        
        # Directly write to the cell
        ws[cell_ref] = update_value
        
        # Save the workbook
        wb.save(file_path)
        return "Success"
    except KeyError:
        return "Failed: Invalid cell reference format or cell does not exist."
    except Exception as e:
        return f"Failed: {str(e)}"

def get_xlsx_last_row(file_path, sheet_name):
    """
    Finds the last row with data and returns its row number and the value in column A.
    
    Args:
        file_path (str): Path to the .xlsx file.
        sheet_name (str): Name of the worksheet.
    
    Returns:
        tuple: (last_row_number, first_column_value) or (None, None) if not found or error.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return None, "Sheet not found"
        ws = wb[sheet_name]
        
        last_row_with_data = None
        # Iterate from the bottom up to find the last row with any data
        # ws.max_row can sometimes be misleading if there are empty formatted rows
        for row_idx in range(ws.max_row, 0, -1):
            row_values = [cell.value for cell in ws[row_idx]]
            if any(v is not None for v in row_values):
                last_row_with_data = row_idx
                break
        
        if last_row_with_data is not None:
            first_col_value = ws.cell(row=last_row_with_data, column=1).value
            return last_row_with_data, first_col_value
        else:
            # Sheet is empty
            return None, None
    except Exception as e:
        return None, str(e)

def get_xlsx_cell_reference_by_value(file_path, sheet_name, cell_value):
    """
    Returns the cell reference (e.g., 'A2', 'C8') for the first cell matching the given value.
    
    Args:
        file_path (str): Path to the .xlsx file.
        sheet_name (str): Name of the worksheet.
        cell_value: Value to search for (case and type sensitive).
    
    Returns:
        str: Cell reference (e.g., 'B5') or an error message.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."
        ws = wb[sheet_name]
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == cell_value:
                    return cell.coordinate # e.g., 'B5'
                    
        return "Error: Value not found in sheet."
    except Exception as e:
        return f"Error: {str(e)}"