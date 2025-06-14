import re
import openpyxl
import datetime
import numbers
from excel_legacy_utils import get_xls_cell_value,update_xls_cell,get_xls_last_row,get_xls_cell_reference_by_value


def clean_number(value: any) -> any:
    """
    Removes the decimal part of a number if it's a whole number (e.g., 45.0 -> 45).
    
    This function safely handles various input types:
    - Floats ending in .0 are converted to integers.
    - Floats with other decimal values are left unchanged.
    - Integers are returned as is.
    - Non-numeric values are returned as is.

    Args:
        value: The input value, which can be a float, int, or any other type.

    Returns:
        An integer if the input was a whole number, otherwise the original value.
    """
    # Check if the value is a number (either int or float)
    if isinstance(value, numbers.Number):
        # A number is whole if its remainder when divided by 1 is 0
        if value % 1 == 0:
            return int(value)
    
    # If it's not a number or has a non-zero decimal, return it as is
    return value

def get_column_values(file_path: str, sheet_name: str, column_name: str) -> list:
    """
    Extracts all values from a specified column in a legacy .xls file.

    This function reads an Excel sheet to find a specific column by its header name,
    then collects all the data in that column from the row after the header to the last
    row containing data in the sheet.

    Args:
        file_path (str): The full path to the .xls file.
        sheet_name (str): The name of the sheet to read from (e.g., "Mass Update").
        column_name (str): The exact name of the column header to find 
                           (e.g., "Shipping Order Number *").

    Returns:
        list: A list of values from the specified column. Returns an empty list
              if the column or sheet is not found or if an error occurs.
    """
    print(f"Starting extraction from '{file_path}'...")
    
    # Find the cell reference of the column header (e.g., 'A15')
    header_cell_ref = get_xls_cell_reference_by_value(file_path, sheet_name, column_name)
    
    if "Error:" in header_cell_ref:
        print(f"Error finding column header: {header_cell_ref}")
        return []
        
    print(f"Found column header '{column_name}' at cell {header_cell_ref}.")

    # Extract the column letter(s) and header row number from the reference
    match = re.match(r"([A-Za-z]+)([0-9]+)", header_cell_ref)
    if not match:
        print("Error: Could not parse the header cell reference.")
        return []
    
    col_letters, header_row_num_str = match.groups()
    start_row = int(header_row_num_str) + 1
    
    # Find the last row with any data in the sheet
    last_row, _ = get_xls_last_row(file_path, sheet_name)
    
    if last_row is None:
        print("Error: Could not determine the last row of the sheet.")
        return []
        
    print(f"Data extraction will run from row {start_row} to {last_row}.")

    # Loop from the row after the header to the last row and collect cell values
    column_values = []
    for current_row in range(start_row, last_row + 1):
        cell_to_read = f"{col_letters}{current_row}"
        value = get_xls_cell_value(file_path, sheet_name, cell_to_read)
        
        # Stop if we hit an error or an empty cell, assuming data is contiguous
        if isinstance(value, str) and "Error:" in value:
            print(f"Encountered an error reading {cell_to_read}: {value}")
            break
        if value == "":
            continue # Skip empty cells

        column_values.append(value)
        
    print("Extraction complete.")
    return column_values

def get_column_values_with_row_numbers(file_path: str, sheet_name: str, column_name: str) -> list:
    """
    Extracts all values and their row numbers from a specified column in a legacy .xls file.

    This function reads an Excel sheet to find a specific column by its header name,
    then collects all the data in that column from the row after the header to the last
    row containing data in the sheet.

    Args:
        file_path (str): The full path to the .xls file.
        sheet_name (str): The name of the sheet to read from (e.g., "Mass Update").
        column_name (str): The exact name of the column header to find 
                           (e.g., "Shipping Order Number *").

    Returns:
        List[Tuple[int, Any]]: A list of tuples, where each tuple contains the row number 
                               and the corresponding cell value (e.g., [(16, 'ValueA'), (17, 'ValueB')]). 
                               Returns an empty list if the column or sheet is not found 
                               or if an error occurs.
    """
    print(f"Starting extraction from '{file_path}'...")
    
    # Find the cell reference of the column header (e.g., 'A15')
    header_cell_ref = get_xls_cell_reference_by_value(file_path, sheet_name, column_name)
    
    if "Error:" in header_cell_ref:
        print(f"Error finding column header: {header_cell_ref}")
        return []
        
    print(f"Found column header '{column_name}' at cell {header_cell_ref}.")

    # Extract the column letter(s) and header row number from the reference
    match = re.match(r"([A-Za-z]+)([0-9]+)", header_cell_ref)
    if not match:
        print("Error: Could not parse the header cell reference.")
        return []
    
    col_letters, header_row_num_str = match.groups()
    start_row = int(header_row_num_str) + 1
    
    # Find the last row with any data in the sheet
    last_row, _ = get_xls_last_row(file_path, sheet_name)
    
    if last_row is None:
        print("Error: Could not determine the last row of the sheet.")
        return []
        
    print(f"Data extraction will run from row {start_row} to {last_row}.")

    # Loop from the row after the header to the last row and collect cell values with their row numbers
    column_data = []
    for current_row in range(start_row, last_row + 1):
        cell_to_read = f"{col_letters}{current_row}"
        value = get_xls_cell_value(file_path, sheet_name, cell_to_read)
        
        # Skip empty cells
        if value == "":
            continue

        # Stop if we hit an error reading a cell
        if isinstance(value, str) and "Error:" in value:
            print(f"Encountered an error reading {cell_to_read}: {value}")
            break

        column_data.append((current_row, value))
        
    print("Extraction complete.")
    return column_data


def find_row_and_get_values(file_path: str, sheet_name: str, search_column_name: str, matching_value: any, columns_to_return: list) -> tuple:
    """
    Finds a row by a specific value in a column and returns the row number
    and a dictionary of values from other specified columns in that row.
    This version dynamically finds the header row and formats dates.

    Args:
        file_path (str): The path to the .xlsx Excel file.
        sheet_name (str): The name of the worksheet to search within.
        search_column_name (str): The header of the column to search for the matching value.
        matching_value: The value to find within the search column.
        columns_to_return (list): A list of column headers whose values should be
                                  returned from the matched row.

    Returns:
        tuple: A tuple containing:
               - The row number (int) where the value was found, or None.
               - A dictionary (dict) with the requested column headers as keys and their
                 corresponding values from the found row. Returns an empty dict if
                 no match is found, or a dict with an 'error' key if an issue occurs.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True, rich_text=True)
        if sheet_name not in workbook.sheetnames:
            return None, {"error": f"Sheet '{sheet_name}' not found."}
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        return None, {"error": f"File not found at path: {file_path}"}
    except Exception as e:
        return None, {"error": f"Failed to open workbook: {e}"}

    header_row_num = None
    headers = {}
    
    for row_num in range(1, min(21, sheet.max_row + 1)):
        row_values = [cell.value for cell in sheet[row_num] if cell.value is not None]
        if search_column_name in row_values:
            header_row_num = row_num
            break

    if header_row_num is None:
        workbook.close()
        return None, {"error": f"Could not find header '{search_column_name}' in the first 20 rows."}

    headers = {cell.value: cell.column for cell in sheet[header_row_num] if cell.value is not None}

    if search_column_name not in headers:
        return None, {"error": f"Search column '{search_column_name}' was found, but failed to map to a column index."}
    for col in columns_to_return:
        if col not in headers:
            workbook.close()
            return None, {"error": f"Column to return '{col}' not found in the header row."}

    search_col_idx = headers[search_column_name]

    for row_index in range(header_row_num + 1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=search_col_idx).value
        
        if cell_value is not None and str(cell_value).strip() == str(matching_value).strip():
            row_data = {}
            for col_name in columns_to_return:
                col_idx = headers[col_name]
                value = sheet.cell(row=row_index, column=col_idx).value
                
                # --- CORRECTED FORMATTING LOGIC ---
                # If the value is a datetime object, format it to a readable string.
                if isinstance(value, datetime.datetime):
                    # This format produces 'YYYY-MM-DD'. You can change it to whatever you need,
                    # for example: '%d-%b-%y' would produce '16-Jun-25'.
                    value = value.strftime('%Y-%m-%d')
                # --- END OF FORMATTING LOGIC ---
                
                row_data[col_name] = value
            
            workbook.close()
            return row_index, row_data

    workbook.close()
    return None, {}